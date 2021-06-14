import sys 
sys.path.append('C:/Daten/cadwork_installation/cadwork.cat/userprofil_28/3d/API.x64/attributes2xl/venv/Lib/site-packages')

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

_file_name = "Test.xlsx"
_sheet_name = "Test_Sheet"

def main():
    new_workbook(_file_name, _sheet_name)

def new_workbook(_file_name, _sheet_name):
    wb = Workbook()  # Workbook Object
    ws = wb.active  # Gets the active worksheet
    ws.title = _sheet_name  # Name the active worksheet

    # Writing the header columns
    ws['A1'] = 'Name'
    ws['B1'] = 'Class'
    ws['C1'] = 'Section'
    ws['D1'] = 'Marks'
    ws['E1'] = 'Age'
    

    col_range = ws.max_column  # get max columns in the worksheet

    # formatting the header columns, filling red color
    for col in range(1, col_range + 1):
        cell_header = ws.cell(1, col)
        cell_header.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid") #used hex code for red color

    wb.save(_file_name)  # save the workbook
    wb.close()  # close the workbook


if __name__ == '__main__':
    main()
    