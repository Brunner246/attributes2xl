import sys
from typing import ClassVar 
sys.path.append('C:/Daten/cadwork_installation/cadwork.cat/userprofil_28/3d/API.x64/attributes2xl/venv/Lib/site-packages')
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import openpyxl

#cell_header = ws.cell(1, col)
#cell_header.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid") #used hex code for red color


class Color():
    """Named colors for use in styles."""
    BLACK = 'FF000000'
    WHITE = 'FFFFFFFF'
    RED = 'FFFF0000'
    DARKRED = 'FF800000'
    BLUE = 'FF0000FF'
    DARKBLUE = 'FF000080'
    GREEN = 'FF00FF00'
    DARKGREEN = 'FF008000'
    YELLOW = 'FFFFFF00'
    DARKYELLOW = 'FF808000'

print(Color.BLACK)

