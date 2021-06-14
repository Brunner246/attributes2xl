

import element_controller as ec, attribute_controller as ac
import sys 
sys.path.append('C:/Daten/cadwork_installation/cadwork.cat/userprofil_28/3d/API.x64/attributes2xl/venv/Lib/site-packages')
import openpyxl
# Call a Workbook() function of openpyxl
# to create a new blank Workbook object
wb = openpyxl.Workbook()

# Get workbook active sheet
# from the active attribute
sheet = wb.active

# Cell objects also have row, column
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or column integer
# is 1, not 0. Cell object is created by
# using sheet object's cell() method.
c1 = sheet.cell(row = 1, column = 1)

# writing values to cells
c1.value = "GUID"

c2 = sheet.cell(row= 1 , column = 2)
c2.value = "my_guid"

# Once have a Worksheet object, one can
# access a cell object by its name also.
# A2 means column = 1 & row = 2.
c3 = sheet['A2']
c3.value = "Name"

# B2 means column = 2 & row = 2.
c4 = sheet['B2']
c4.value = "my_name"

c4 = sheet['A3']
c4.value = "Baugruppe"

c5 = sheet['B4']
c5.value = 'Obergeschoss'


# Anytime you modify the Workbook object
# or its sheets and cells, the spreadsheet
# file will not be saved until you call
# the save() workbook method.
wb.save("demo.xlsx")
